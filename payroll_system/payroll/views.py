from django.shortcuts import get_object_or_404, redirect, render
from .forms import PayrollForm,PayrollUploadForm
from .models import Payroll
from django.contrib import messages
from django.utils.dateparse import parse_datetime
from employee.models import Employee
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from openpyxl import Workbook
from django.contrib.staticfiles import finders
from django.db.models import Min, Max
from django.utils import timezone
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from django.utils.dateparse import parse_date
from django.db import IntegrityError
import pandas as pd
from datetime import datetime
from django.core.exceptions import ValidationError
from decimal import Decimal

def dashboard(request):
    return redirect('employee_list')

def generatePayroll(request):
    """
    View to generate payroll for an employee.

    The form collects details such as time in, time out, total hours worked, deductions,
    overtime pay, and night differential pay. It computes the gross salary, net salary, 
    and the final payroll for the employee. The result is saved to the Payroll model.

    - POST method: Handles payroll generation and saves the data.
    - GET method: Displays an empty payroll form.
    """
    if request.method == 'POST':
        form = PayrollForm(request.POST)

        if form.is_valid():
            try:
                employee = form.cleaned_data['employee']
                time_in = form.cleaned_data['time_in']
                time_out = form.cleaned_data['time_out']

                if isinstance(time_in, str) and time_in:
                    time_in = parse_datetime(time_in)
                if isinstance(time_out, str) and time_out:
                    time_out = parse_datetime(time_out)

                total_hours_worked = form.cleaned_data['total_hours_worked']
                deductions = form.cleaned_data['deductions']
                subtotal = form.cleaned_data['subtotal']
                net_salary = form.cleaned_data['net_salary']
                deduction_remarks = form.cleaned_data['deduction_remarks']
                project = form.cleaned_data['project']
                daily_rate = form.cleaned_data['daily_rate']
                overtime_pay = form.cleaned_data['overtime_pay']
                overtime_hour = form.cleaned_data['overtime_hour']
                night_differential_pay = form.cleaned_data['night_differential_pay']
                night_differential_hour = form.cleaned_data['night_differential_hour']
                allowance = form.cleaned_data['allowance']
                date = time_in.date()
            
                payroll = Payroll(
                    employee=employee,
                    time_in=time_in,
                    time_out=time_out,
                    total_hours_worked=total_hours_worked,
                    deductions=deductions,
                    subtotal=subtotal,
                    net_salary=net_salary,
                    deduction_remarks=deduction_remarks,
                    project=project,
                    daily_rate=daily_rate,
                    overtime_pay = overtime_pay,
                    overtime_hour = overtime_hour,
                    night_differential_pay = night_differential_pay,
                    night_differential_hour = night_differential_hour,  
                    allowance = allowance,
                    date=date
                )
                print("valid",payroll)
                payroll.save()

                messages.success(request, f"Payroll for {employee} has been successfully generated!")
                form = PayrollForm()
                return render(request,'generate_payroll.html', {'form': form})
            except IntegrityError:
                print("error")
                messages.error(request, 'A payroll record for this employee on this date already exists.')
                return render(request, 'generate_payroll.html', {
                    'form': form,
                    'total_hours_worked': form.cleaned_data.get('total_hours_worked', 0),
                })
        else:
            for field in form:
                for error in field.errors:
                    messages.error(request, f"Error in {field.label}: {error}")
            return render(request, 'generate_payroll.html', {
                'form': form,
                'total_hours_worked': form.cleaned_data.get('total_hours_worked', 0),
            })

    else:
        form = PayrollForm()
        return render(request, 'generate_payroll.html', {'form': form})

def payrollSummary(request):
    """
    View to display the payroll summary for all employees or a specific employee.

    This view displays a list of generated payrolls, with the option to filter by employee.
    The summary will show the payroll details such as the employee's name, pay period, 
    gross salary, deductions, and net salary.
    """
    employees = Employee.objects.all()
    selected_employee = None
    payrolls = Payroll.objects.all()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if request.method == 'GET':
        selected_employee_id = request.GET.get('employee')
        if selected_employee_id:
            selected_employee = get_object_or_404(Employee, id=selected_employee_id)
            payrolls = Payroll.objects.filter(employee=selected_employee)
        if start_date:
            payrolls = payrolls.filter(date__gte=parse_date(start_date))
        if end_date:
            payrolls = payrolls.filter(date__lte=parse_date(end_date))

        total_hours_worked = sum(payroll.total_hours_worked for payroll in payrolls)
        total_overtime_pay = sum(payroll.overtime_pay for payroll in payrolls)
        total_night_differential_pay = sum(payroll.night_differential_pay for payroll in payrolls)
        allowance = sum(payroll.allowance for payroll in payrolls)
        total_deductions = sum(payroll.deductions for payroll in payrolls)
        total_gross_salary = sum(payroll.subtotal for payroll in payrolls)
        total_net_salary = sum(payroll.net_salary for payroll in payrolls)

        return render(request, 'payroll_summary.html', {
            'payrolls': payrolls,
            'employees': employees,
            'selected_employee': selected_employee,
            'selected_employee_id': selected_employee_id,
            'start_date': start_date,
            'end_date': end_date,
            'total_hours_worked': total_hours_worked,
            'total_overtime_pay': total_overtime_pay,
            'total_night_differential_pay': total_night_differential_pay,
            'allowance': allowance,
            'total_deductions': total_deductions,
            'total_gross_salary': total_gross_salary,
            'total_net_salary': total_net_salary
        })

def exportPayslipPdf(request, employee_id):
    """
    Generate and return a PDF payslip for a specific employee, including the company logo,
    name, position, pay period, date generated, daily rate, and summary.
    """
    employee = get_object_or_404(Employee, id=employee_id)
    payrolls = Payroll.objects.filter(employee=employee)

    if not payrolls:
        return HttpResponse("No payroll records found for this employee.", status=404)

    total_hours_worked = sum(payroll.total_hours_worked for payroll in payrolls)
    total_overtime_pay = sum(payroll.overtime_pay for payroll in payrolls)
    total_night_differential_pay = sum(payroll.night_differential_pay for payroll in payrolls)
    allowance = sum(payroll.allowance for payroll in payrolls)
    total_deductions = sum(payroll.deductions for payroll in payrolls)
    total_gross_salary = sum(payroll.subtotal for payroll in payrolls)
    total_net_salary = sum(payroll.net_salary for payroll in payrolls)
    pay_period_from = payrolls.aggregate(Min('date'))['date__min']
    pay_period_to = payrolls.aggregate(Max('date'))['date__max']
    current_date = timezone.now()
    daily_rate = payrolls.first().daily_rate

    html_string = render_to_string('payroll_payslip.html', {
        'selected_employee': employee,
        'payrolls': payrolls,
        'total_hours_worked': total_hours_worked,
        'total_overtime_pay': total_overtime_pay,
        'total_night_differential_pay': total_night_differential_pay,
        'allowance': allowance,
        'total_deductions': total_deductions,
        'total_gross_salary': total_gross_salary,
        'total_net_salary': total_net_salary,
        'pay_period_from': pay_period_from,
        'pay_period_to': pay_period_to,
        'current_date': current_date,
        'daily_rate': daily_rate
    })

    css_path = finders.find('css/payslip.css')

    if not css_path:
        return HttpResponse("CSS file not found.", status=404)

    pdf_file = HTML(string=html_string,base_url=request.build_absolute_uri()).write_pdf(stylesheets=[css_path])

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payslip_{employee.first_name}_{employee.last_name}.pdf"'
    return response


def generatePayslipExcel(request, employee_id):
    """
    Generate and return an Excel payslip for a specific employee, including the company logo, 
    name, position, pay period, date generated, daily rate, and summary.
    """
    employee = get_object_or_404(Employee, id=employee_id)
    payrolls = Payroll.objects.filter(employee=employee).select_related('employee')

    if not payrolls:
        return HttpResponse("No payroll records found for this employee.", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payslip"

    logo_path = finders.find('img/company_logo.png')
    if logo_path:
        img = Image(logo_path)
        img.width = 120
        img.height = 80
        ws.add_image(img, 'A1')
    
    ws.merge_cells('C1:G1')
    ws['C1'] = "HODREAL FIT-OUT AND CONSTRUCTION"
    ws['C1'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C1'].font = Font(bold=True, size=14)
    
    ws.merge_cells('C2:J2')
    ws['C2'] = "Bantangas City | Contact: 09217292222 | Email: hodrealconstruction@yahoo.com"
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C2'].font = Font(size=10)

    row_offset = 5 

    ws.merge_cells(f'A{row_offset}:F{row_offset}')
    ws[f'A{row_offset}'] = f"Employee: {employee.first_name} {employee.last_name}"
    ws[f'A{row_offset}'].font = Font(bold=True)
    ws[f'A{row_offset}'].alignment = Alignment(horizontal='left')

    ws.merge_cells(f'A{row_offset + 1}:F{row_offset + 1}')
    ws[f'A{row_offset + 1}'] = f"Position: {employee.position}"
    ws[f'A{row_offset + 1}'].font = Font(bold=True)
    ws[f'A{row_offset + 1}'].alignment = Alignment(horizontal='left')

    pay_period_from = payrolls.aggregate(Min('date'))['date__min']
    pay_period_to = payrolls.aggregate(Max('date'))['date__max']

    ws.merge_cells(f'A{row_offset + 2}:F{row_offset + 2}')
    ws[f'A{row_offset + 2}'] = f"Pay Period: {pay_period_from.strftime('%Y-%m-%d')} - {pay_period_to.strftime('%Y-%m-%d')}"
    ws[f'A{row_offset + 2}'].font = Font(bold=True)
    ws[f'A{row_offset + 2}'].alignment = Alignment(horizontal='left')

    ws.merge_cells(f'A{row_offset + 3}:F{row_offset + 3}')
    ws[f'A{row_offset + 3}'] = f"Date Generated: {timezone.now().strftime('%Y-%m-%d')}"
    ws[f'A{row_offset + 3}'].font = Font(bold=True)
    ws[f'A{row_offset + 3}'].alignment = Alignment(horizontal='left')

    daily_rate = payrolls.first().daily_rate
    ws.merge_cells(f'A{row_offset + 4}:F{row_offset + 4}')
    ws[f'A{row_offset + 4}'] = f"Daily Rate: {daily_rate}"
    ws[f'A{row_offset + 4}'].font = Font(bold=True)
    ws[f'A{row_offset + 4}'].alignment = Alignment(horizontal='left')

    row_offset += 7

    headers = [
        "Date", "Overtime", "Night Differential", "Allowance", "Deductions", "Total Amount"
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=row_offset, column=col_num, value=header)
    
    for payroll in payrolls:
        ws.append([
            payroll.date.strftime("%Y-%m-%d"),
            payroll.overtime_pay,
            payroll.night_differential_pay,
            payroll.allowance,
            payroll.deductions,
            payroll.net_salary
        ])

    total_hours_worked = sum(payroll.total_hours_worked for payroll in payrolls)
    total_overtime_pay = sum(payroll.overtime_pay for payroll in payrolls)
    total_night_differential_pay = sum(payroll.night_differential_pay for payroll in payrolls)
    total_allowance = sum(payroll.allowance for payroll in payrolls)
    total_deductions = sum(payroll.deductions for payroll in payrolls)
    total_gross_salary = sum(payroll.subtotal for payroll in payrolls)
    total_net_salary = sum(payroll.net_salary for payroll in payrolls)

    row_offset += len(payrolls) + 3 

    ws[f'A{row_offset}'] = f"Total Hours Worked: {total_hours_worked}"
    ws[f'A{row_offset + 1}'] = f"Total Overtime Pay: {total_overtime_pay}"
    ws[f'A{row_offset + 2}'] = f"Total Night Differential Pay: {total_night_differential_pay}"
    ws[f'A{row_offset + 3}'] = f"Total Allowance: {total_allowance}"
    ws[f'A{row_offset + 4}'] = f"Total Deductions: {total_deductions}"
    ws[f'A{row_offset + 5}'] = f"Total Gross Salary: {total_gross_salary}"
    ws[f'A{row_offset + 6}'] = f"Total Net Salary: {total_net_salary}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payslip_{employee.first_name}_{employee.last_name}.xlsx"'

    wb.save(response)
    return response

def editPayroll(request, payroll_id):
    """
    Edits an existing payroll record based on the provided payroll ID.

    This view handles both displaying the edit form and processing the submitted data.
    - If the request method is GET, the current payroll details are displayed in a form for editing.
    - If the request method is POST, the form is validated, and if successful, the payroll record is updated in the database.
    - In case of a successful update, a success message is displayed, and the user is redirected to the payroll summary page.
    - If there is an error while updating the record (e.g., database constraint issues), an error message is shown to the user.
    - If the form contains errors, those errors are displayed next to the relevant fields.

    Args:
        request (HttpRequest): The HTTP request object.
        payroll_id (int): The ID of the payroll record to be edited.

    Returns:
        HttpResponse: Renders the 'edit_payroll.html' template with the form and payroll data.
        If the form is valid, the user is redirected to the payroll summary page.
    """
    payroll = get_object_or_404(Payroll, id=payroll_id)
    if request.method == 'POST':
        form = PayrollForm(request.POST, instance=payroll)
        if form.is_valid():
            try:
                time_in = form.cleaned_data['time_in']
                if isinstance(time_in, str) and time_in:
                    time_in = parse_datetime(time_in)
                payroll.date = time_in.date()
                form.save()
                messages.success(request, 'Payroll record updated successfully.')
                return redirect('payroll_summary')
            except IntegrityError as e:
                messages.error(request, f'Error updating payroll record: {e}')
        else:
            # print(form.errors)
            for field in form:
                for error in field.errors:
                    messages.error(request, f"Error in {field.label}: {error}")
            return render(request, 'edit_payroll.html', {
                'form': form,
                'payroll': payroll
            })
    else:
        form = PayrollForm(instance=payroll)
        # print(form.initial)
    return render(request, 'edit_payroll.html', {'form': form})

def deletePayroll(request, payroll_id):
    """
    Deletes a payroll record based on the provided payroll ID.

    This view handles the deletion of an existing payroll record from the database.
    - If the request method is POST, the payroll record is deleted, and a success message is displayed.
    - After deletion, the user is redirected to the payroll summary page.

    Args:
        request (HttpRequest): The HTTP request object.
        payroll_id (int): The ID of the payroll record to be deleted.

    Returns:
        HttpResponse: Redirects to the 'payroll_summary' page after the payroll record is deleted.
    """
    payroll = get_object_or_404(Payroll, id=payroll_id)
    if request.method == 'POST':
        payroll.delete()
        messages.success(request, 'Payroll record deleted successfully.')
    return redirect('payroll_summary')

def batchUpload(request):
    """
    Handles the bulk upload of payroll records from an Excel file.

    This view allows users to upload an Excel file containing payroll information, validate the data, 
    and save the records to the database. The process includes the following steps:
    - The user uploads an Excel file containing payroll data.
    - The uploaded file is validated to ensure it contains the required columns.
    - Missing or invalid data is handled appropriately (e.g., setting missing values to defaults).
    - The payroll records are created for each row in the uploaded file, associating each record with an existing employee.
    - Errors in processing, such as missing required columns or invalid employee IDs, are communicated to the user.

    Args:
        request (HttpRequest): The HTTP request object.
    Returns:
        HttpResponse: Renders the 'batch_upload.html' template with the form. 
            If the file is successfully uploaded, the user is redirected with a success message.
            If there are errors during the process (e.g., missing columns, employee not found), 
            an error message is shown and the user is redirected to the upload page.
    """
    # Date format: yyyy-mm-dd
    # Time format: hh:mm AM/PM
    if request.method == 'POST' and request.FILES['excel_file']:
        form = PayrollUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                df = pd.read_excel(excel_file)
                required_columns = [
                    'employee_id', 'daily_rate', 'allowance', 'total_hours_worked', 'overtime_pay', 
                    'overtime_hour', 'night_differential_pay', 'night_differential_hour', 
                    'deductions', 'deduction_remarks', 'subtotal', 'net_salary', 'date', 
                    'time_in', 'time_out', 'project'
                ]
                for column in required_columns:
                    if column not in df.columns:
                        messages.error(request, f'Missing required column: {column}')
                        return redirect('payroll_batch_upload')
                    
                # df = df.applymap(lambda x: '' if pd.isna(x) else x)
                for column in df.columns:
                    for index, value in df[column].items():
                        if pd.isna(value):
                            field_type = Payroll._meta.get_field(column).get_internal_type()
                            if field_type == 'DecimalField':
                                df.at[index, column] = Decimal('0.00') 
                            elif field_type == 'CharField':
                                df.at[index, column] = '' 
                payroll_list = []
                for index, row in df.iterrows():
                    try:
                        employee = Employee.objects.get(id=row['employee_id'])
                        time_in = pd.to_datetime(row['time_in'], format='%H:%M:%S').time()
                        time_out = pd.to_datetime(row['time_out'], format='%H:%M:%S').time()
                        date_value = pd.to_datetime(row['date']).date()
                        time_in_full = datetime.combine(date_value, time_in)
                        time_out_full = datetime.combine(date_value, time_out)
                        payroll = Payroll(
                            employee=employee,
                            daily_rate=row['daily_rate'],
                            allowance=row['allowance'],
                            total_hours_worked=row['total_hours_worked'],
                            overtime_pay=row['overtime_pay'],
                            overtime_hour=row['overtime_hour'],
                            night_differential_pay=row['night_differential_pay'],
                            night_differential_hour=row['night_differential_hour'],
                            deductions=row['deductions'],
                            deduction_remarks=row.get('deduction_remarks', ''),
                            subtotal=row['subtotal'],
                            net_salary=row['net_salary'],
                            date=pd.to_datetime(row['date']).date(),
                            time_in=time_in_full,
                            time_out=time_out_full,
                            project=row.get('project', ''),
                        )
                        payroll_list.append(payroll)
                    except Employee.DoesNotExist:
                        messages.error(request, f"Employee with ID {row['employee_id']} does not exist.")
                        return redirect('payroll_batch_upload')
                    except Exception as e:
                        messages.error(request, f"An error occurred: {str(e)}")
                        return redirect('payroll_batch_upload')

                Payroll.objects.bulk_create(payroll_list)
                messages.success(request, 'Payroll records uploaded successfully!')
                return redirect('payroll_batch_upload')
            except IntegrityError as e:
                messages.error(request, f'Error uploading payroll record: Duplicate record for date entry.')
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                return redirect('payroll_batch_upload')
    else:
        form = PayrollUploadForm()

    return render(request, 'batch_upload.html', {'form': form})

def downloadTemplate(request):
    """
    Generates and downloads an Excel template for payroll data entry.

    This view creates an Excel file with predefined column headers and blank fields, 
    which serves as a template for users to fill in payroll data. The template includes columns 
    such as employee ID, daily rate, allowances, total hours worked, overtime pay, and more. 
    The generated file is returned as an Excel file for download.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: An Excel file containing the payroll template, which the user can download.
                      The file is named 'payroll_template.xlsx' and includes predefined columns 
                      with empty values for data entry.
    """
    columns = [
        'employee_id', 'daily_rate', 'allowance', 'total_hours_worked', 'overtime_pay', 
        'overtime_hour', 'night_differential_pay', 'night_differential_hour', 
        'deductions', 'deduction_remarks', 'subtotal', 'net_salary', 'date', 
        'time_in', 'time_out', 'project'
    ]
    data = {
        'employee_id': [''],
        'daily_rate': [''],
        'allowance': [0],
        'total_hours_worked': [0],
        'overtime_pay': [0],
        'overtime_hour': [0],
        'night_differential_pay': [0],
        'night_differential_hour': [0],
        'deductions': [0],
        'deduction_remarks': [''],
        'subtotal': [0],
        'net_salary': [0],
        'date': [''],
        'time_in': [''],
        'time_out': [''],
        'project': ['']
    }
    df = pd.DataFrame(data, columns=columns)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payroll_template.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Payroll Template')

    return response
